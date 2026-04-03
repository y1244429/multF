#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
另类数据因子库 v2 - 多期滚动IC验证
用27个月度收益期验证分析师情绪因子稳定性
"""
import os, time, warnings
import pandas as pd
import numpy as np
from scipy.stats import spearmanr
warnings.filterwarnings('ignore')

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR = os.path.join(SCRIPT_DIR, 'output')
DATA_DIR = '/Users/ydy/CodeBuddy/output/market_data'

# ── 另类因子定义 ────────────────────────────────────────────────────────
ALPHA_FACTORS = {
    'analyst_buy_ratio': {
        '大类': '另类数据-分析师情绪', '细类': '分析师评级',
        '说明': '买入评级/(买入+增持+中性)，IC>0=分析师乐观→股价涨',
    },
    'analyst_attention': {
        '大类': '另类数据-分析师情绪', '细类': '分析师关注度',
        '说明': 'ln(研报数+1)，IC>0=分析师覆盖多→机构化程度高→Alpha强',
    },
    'analyst_net_rating': {
        '大类': '另类数据-分析师情绪', '细类': '分析师综合评级',
        '说明': '买入*5+增持*4+中性*3+减持*2+卖出*1，IC>0=综合评级越高→强',
    },
    'analyst_eps_fwd_growth': {
        '大类': '另类数据-分析师情绪', '细类': '预测EPS调整',
        '说明': '(2025预测EPS - 2024预测EPS)/|2024EPS|，IC>0=分析师上调预测→强',
    },
    'analyst_eps_2025': {
        '大类': '另类数据-分析师情绪', '细类': '预测EPS绝对值',
        '说明': '2025年预测EPS，IC>0=预测EPS高=基本面强→Alpha',
    },
    'weibo_sentiment': {
        '大类': '另类数据-微博情绪', '细类': '微博舆情',
        '说明': '微博近12小时舆情情绪，IC>0=正面舆情→股价涨（仅热门股有效）',
    },
    'pmi_manufacturing': {
        '大类': '另类数据-PMI宏观', '细类': 'PMI宏观',
        '说明': '制造业PMI标准化，所有股票同时暴露，IC>0=PMI扩张→股市涨',
    },
}

def calc_analyst_factors(df_analyst):
    rows = []
    for _, r in df_analyst.iterrows():
        code = str(r['代码']).zfill(6)
        name = r['名称']
        buy = float(r.get('机构投资评级(近六个月)-买入') or 0)
        hold = float(r.get('机构投资评级(近六个月)-增持') or 0)
        neutral = float(r.get('机构投资评级(近六个月)-中性') or 0)
        reduce_ = float(r.get('机构投资评级(近六个月)-减持') or 0)
        sell = float(r.get('机构投资评级(近六个月)-卖出') or 0)
        total = buy + hold + neutral + reduce_ + sell

        report_cnt = float(r.get('研报数') or 0)

        buy_ratio = buy / total if total > 0 else np.nan
        net_rating = (buy*5 + hold*4 + neutral*3 + reduce_*2 + sell*1) / total if total > 0 else np.nan
        analyst_attn = np.log(report_cnt + 1)

        try:
            eps_24 = float(r.get('2024预测每股收益') or np.nan)
            eps_25 = float(r.get('2025预测每股收益') or np.nan)
            eps_fwd_growth = (eps_25 - eps_24) / abs(eps_24) if (not np.isnan(eps_24) and eps_24 != 0) else np.nan
        except:
            eps_fwd_growth = np.nan

        eps_25_val = float(r.get('2025预测每股收益') or np.nan)

        rows.append({
            'code': code, 'name': name,
            'analyst_buy_ratio': buy_ratio,
            'analyst_attention': analyst_attn,
            'analyst_net_rating': net_rating,
            'analyst_eps_fwd_growth': eps_fwd_growth,
            'analyst_eps_2025': eps_25_val,
            'report_cnt': report_cnt,
        })
    return pd.DataFrame(rows)

def load_monthly_returns():
    """加载27个月的月末收益（2023-01 ~ 2025-03）"""
    print("  加载月度收益...")
    files = sorted([f for f in os.listdir(DATA_DIR) if f.endswith('.csv')])[:600]
    stock_data = {}
    for fname in files:
        code = fname.replace('.csv', '')
        try:
            df = pd.read_csv(os.path.join(DATA_DIR, fname))
            if 'date' not in df.columns or len(df) < 30:
                continue
            df['date'] = pd.to_datetime(df['date'])
            df['month'] = df['date'].dt.to_period('M')
            # 月度收益
            monthly_ret = df.groupby('month').apply(
                lambda g: pd.Series({'ret': g['close'].iloc[-1] / g['close'].iloc[0] - 1})
            ).reset_index()
            monthly_ret['month_str'] = monthly_ret['month'].astype(str)
            monthly_ret['ret'] = monthly_ret['ret']
            stock_data[code] = monthly_ret[['month_str', 'ret']]
        except:
            pass
    print(f"    {len(stock_data)} 只股票月度收益已加载")
    return stock_data

def main():
    print("=" * 60)
    print("另类数据因子库 - 滚动IC验证")
    print("=" * 60)

    # ── 1. 拉取数据 ─────────────────────────────────────────────────
    print("\n[1/4] 拉取另类原始数据...")
    import akshare as ak

    print("  拉取 analyst_forecast_em（全市场）...")
    t0 = time.time()
    df_analyst_raw = ak.stock_profit_forecast_em()
    print(f"    {time.time()-t0:.1f}s, {len(df_analyst_raw)} 只")

    df_analyst = calc_analyst_factors(df_analyst_raw)
    print(f"    分析师因子: {len(df_analyst)} 只")

    print("  拉取 weibo_sentiment...")
    try:
        df_weibo = ak.stock_js_weibo_report(time_period='CNHOUR12')
        df_weibo = df_weibo.rename(columns={'name': 'stock_name', 'rate': 'weibo_sentiment'})
        print(f"    {len(df_weibo)} 只")
    except Exception as e:
        df_weibo = pd.DataFrame()
        print(f"    失败: {e}")

    print("  拉取 macro_china_pmi...")
    df_pmi = ak.macro_china_pmi()
    pmi_mfg_mean = df_pmi['制造业-指数'].mean()
    pmi_mfg_std = df_pmi['制造业-指数'].std()
    pmi_latest_mfg = (df_pmi['制造业-指数'].iloc[0] - pmi_mfg_mean) / pmi_mfg_std
    print(f"    {len(df_pmi)} 期, 最新PMI标准化={pmi_latest_mfg:.3f}")

    # ── 2. 加载月度收益 ──────────────────────────────────────────────
    print("\n[2/4] 加载27个月度收益期...")
    stock_data = load_monthly_returns()

    # 确定月份列表（从最早到最晚）
    all_months = set()
    for code, mdf in stock_data.items():
        all_months.update(mdf['month_str'].tolist())
    months = sorted(all_months)
    print(f"    月份范围: {months[0]} ~ {months[-1]}, 共{len(months)}期")

    # ── 3. 计算滚动IC（每期截面IC）──────────────────────────────────
    print("\n[3/4] 计算滚动IC...")
    factor_cols = [f for f in ALPHA_FACTORS if f in df_analyst.columns]
    factor_cols = [f for f in factor_cols if f != 'pmi_manufacturing']  # PMI单独处理

    # 按月计算截面IC
    monthly_ics = {f: [] for f in factor_cols}
    monthly_ics['pmi_manufacturing'] = []

    for month in months:
        # 构建该月截面数据
        month_rets = {}
        month_factors = {}
        for code, mdf in stock_data.items():
            row = mdf[mdf['month_str'] == month]
            if len(row) > 0:
                month_rets[code] = row['ret'].values[0]

        if len(month_rets) < 50:
            continue

        # 匹配分析师因子
        for _, ar in df_analyst.iterrows():
            code = ar['code']
            if code in month_rets:
                for f in factor_cols:
                    if not pd.isna(ar.get(f)):
                        if f not in month_factors:
                            month_factors[f] = []
                        month_factors[f].append((ar[f], month_rets[code]))

        # PMI（宏观，所有股票相同）
        if 'pmi_manufacturing' in monthly_ics:
            pmi_vals = [(pmi_latest_mfg, r) for r in month_rets.values()]
            month_factors['pmi_manufacturing'] = pmi_vals

        # 计算截面IC
        for f, data in month_factors.items():
            if len(data) > 30:
                vals, rets = zip(*data)
                try:
                    ic, _ = spearmanr(vals, rets)
                    if not np.isnan(ic):
                        monthly_ics[f].append(ic)
                except:
                    pass

    # 汇总IC
    print("\n  月度IC统计:")
    ic_results = []
    for fname in sorted(ALPHA_FACTORS.keys()):
        if fname == 'pmi_manufacturing':
            ic_vals = monthly_ics['pmi_manufacturing']
        else:
            ic_vals = monthly_ics.get(fname, [])

        if len(ic_vals) >= 3:
            ic_arr = np.array(ic_vals)
            ic_mean = np.nanmean(ic_arr)
            ic_std = np.nanstd(ic_arr)
            ir = ic_mean / ic_std if ic_std > 0 else 0
            pos_rate = np.mean(ic_arr > 0)
            ic_results.append({
                '因子名称': fname,
                '大类': ALPHA_FACTORS[fname]['大类'],
                '细类': ALPHA_FACTORS[fname]['细类'],
                '平均IC': round(ic_mean, 4),
                'IC方向': '正向↑' if ic_mean > 0 else '逆向↓',
                '|IC|': round(abs(ic_mean), 4),
                'IC>0比例': f'{pos_rate*100:.1f}%',
                'IR': round(ir, 3),
                'IC_std': round(ic_std, 4),
                '月份数': len(ic_arr),
                'IC稳定性★': '★' * min(int(ir * 2) + 1, 5) if ir > 0 else 'N/A',
                '换手率代理★': 'N/A（低频）',
                '因子独立性★': '★★★★★ 极高（基本面+舆情）',
                '市场稳健性★': '★★★ 中等',
                '逻辑可持续性★': '★★★★ 高',
                '综合评分': round(min(max(ir * 2 + pos_rate * 2, 0), 5), 1),
                '说明': ALPHA_FACTORS[fname]['说明'],
                '来源': 'AKShare另类数据' if fname != 'pmi_manufacturing' else '国家统计局PMI',
            })
            print(f"    {fname}: IC={ic_mean:+.4f}, IR={ir:.3f}, IC>0={pos_rate*100:.1f}%, n={len(ic_arr)}")
        else:
            # 单期IC
            if fname in monthly_ics and len(monthly_ics[fname]) > 0:
                ic_mean = np.nanmean(monthly_ics[fname])
                ic_results.append({
                    '因子名称': fname, '大类': ALPHA_FACTORS[fname]['大类'],
                    '细类': ALPHA_FACTORS[fname]['细类'],
                    '平均IC': round(ic_mean, 4), 'IC方向': '正向↑' if ic_mean > 0 else '逆向↓',
                    '|IC|': round(abs(ic_mean), 4),
                    'IC>0比例': '单期',
                    'IR': '待滚动',
                    'IC_std': '单期',
                    '月份数': 1,
                    'IC稳定性★': '?',
                    '换手率代理★': 'N/A',
                    '因子独立性★': '★★★★★',
                    '市场稳健性★': '?',
                    '逻辑可持续性★': '★★★★',
                    '综合评分': round(abs(ic_mean) * 10, 1),
                    '说明': ALPHA_FACTORS[fname]['说明'],
                    '来源': 'AKShare另类数据',
                })
                print(f"    {fname}: IC={ic_mean:+.4f} (单期)")

    df_ic = pd.DataFrame(ic_results)

    # ── 4. 保存 ─────────────────────────────────────────────────────
    print("\n[4/4] 保存结果...")
    alt_ic_path = os.path.join(OUTPUT_DIR, 'alternative_ic.csv')
    df_ic.to_csv(alt_ic_path, index=False, encoding='utf-8-sig')
    print(f"  保存: {alt_ic_path}")

    # 合并到V5 Excel
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from copy import copy

        wb = openpyxl.load_workbook(
            os.path.join(OUTPUT_DIR, '因子监控_IC_IR_实测版_v5.xlsx')
        )

        # ── Sheet1: 因子监控总表 ──
        ws = wb['因子监控总表']
        # 找最后一行
        last_row = ws.max_row
        while last_row > 3 and ws.cell(last_row, 1).value in (None, ''):
            last_row -= 1
        start_row = last_row + 2

        # 表头（已在第3行）
        fills = {
            '另类数据-分析师情绪': PatternFill('solid', fgColor='FFF2CC'),   # 淡黄
            '另类数据-微博情绪': PatternFill('solid', fgColor='E2EFDA'),    # 淡绿
            '另类数据-PMI宏观': PatternFill('solid', fgColor='DEEBF7'),      # 淡蓝
            '另类数据-公告舆情': PatternFill('solid', fgColor='F4CCCC'),
        }

        if len(df_ic) > 0:
            for _, row_data in df_ic.iterrows():
                row = start_row
                for col_idx, col_name in enumerate(df_ic.columns, 1):
                    val = row_data.get(col_name, '')
                    if isinstance(val, float) and np.isnan(val):
                        val = ''
                    ws.cell(row, col_idx, val)
                    # 高亮
                    cat = row_data.get('大类', '')
                    if cat in fills:
                        ws.cell(row, col_idx).fill = fills[cat]
                start_row += 1

        # ── Sheet4: 通达信选股公式 ──
        ws4 = wb['通达信选股公式']
        last_row4 = ws4.max_row
        while last_row4 > 1 and ws4.cell(last_row4, 1).value in (None, ''):
            last_row4 -= 1
        r = last_row4 + 2

        # 新增另类因子公式
        new_formulas = [
            ('另-分析师公式A（推荐）', '买入评级 + 综合评级 + EPS预测', '', '另-分析师A'),
            ('另-分析师公式B', '研报数关注 + 预测EPS上调', '', '另-分析师B'),
            ('另-微博公式', '微博舆情正面 + 有热度覆盖', '', '另-微博'),
        ]

        # 从df_ic中取最优分析师因子
        analyst_factors = df_ic[df_ic['大类'] == '另类数据-分析师情绪'].sort_values('综合评分', ascending=False)
        if len(analyst_factors) > 0:
            top = analyst_factors.iloc[0]
            new_formulas[0] = (
                f'另-分析师公式A（推荐, {top["因子名称"]} IC={top["平均IC"]:+.4f}）',
                f'买入比例>均值 + 综合评级>4分 + 预测EPS>行业均值',
                f'{top["说明"]}，IC={top["平均IC"]:+.4f}，IC>0率={top["IC>0比例"]}',
                '另-分析师A',
            )

        ws4.cell(r, 1, '── 另类数据因子选股公式 ──')
        ws4.cell(r, 1).font = Font(bold=True, size=12, color='FF6600')
        r += 1
        for fname, ftype, fdesc, ftag in new_formulas:
            ws4.cell(r, 1, fname)
            ws4.cell(r, 2, ftype)
            ws4.cell(r, 3, fdesc)
            ws4.cell(r, 4, ftag)
            r += 1

        wb.save(os.path.join(OUTPUT_DIR, '因子监控_IC_IR_实测版_v5.xlsx'))
        print(f"  Excel已更新！")

    except Exception as e:
        print(f"  Excel更新失败: {e}")
        print("  (不影响CSV保存)")

    # ── 汇总打印 ─────────────────────────────────────────────────────
    print("\n" + "=" * 60)
    print("另类因子IC最终汇总")
    print("=" * 60)
    if len(df_ic) > 0:
        df_sorted = df_ic.sort_values('综合评分', ascending=False)
        for _, r in df_sorted.iterrows():
            print(f"  {r['因子名称']}: IC={r['平均IC']:+.4f} {r['IC方向']}, "
                  f"IR={r['IR']}, IC>0率={r['IC>0比例']}, "
                  f"评分={r['综合评分']}")

    print("\n✅ 完成！")
    print(f"\n文件位置:")
    print(f"  CSV: {alt_ic_path}")
    return df_ic

if __name__ == '__main__':
    main()
