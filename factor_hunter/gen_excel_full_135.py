#!/usr/bin/env python3
"""
合并三类因子：价量(111) + 基本面(19) + 另类(5) = 135因子统一IC监控
参考V6 Excel格式，生成完整监控Excel
"""
import os, warnings, math
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.chart import BarChart, Reference
warnings.filterwarnings('ignore')

OUTPUT_DIR = '/Users/ydy/WorkBuddy/20260326134350/factor_hunter/output'
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ── 颜色定义 ──────────────────────────────────────────────────────────────────
C_HEADER_BG   = '1F4E79'   # 深蓝表头
C_HEADER_FG   = 'FFFFFF'
C_SUB_BG      = '2E75B6'   # 次级蓝
C_ALT_ROW     = 'D6E4F0'   # 浅蓝斑马
C_WHITE       = 'FFFFFF'
C_GOLD        = 'FFD700'
C_STAR_FILL   = 'FFC000'

# 大类颜色
CAT_COLORS = {
    '价量技术类':    '4472C4',
    'AlphaForge':   '70AD47',
    'AlphaGPS':     'ED7D31',
    'AlphaGen':     'A9D18E',
    'HRFT':         'FF0000',
    'QFR':          '9DC3E6',
    'AlphaSAGE':    'F4B183',
    'MASTER':       'C9C9C9',
    '清华IGSM':     'FFC000',
    '北大光华':     'FF6699',
    '基本面-价值':   '7030A0',
    '基本面-成长':   '00B050',
    '基本面-质量':   '00B0F0',
    '基本面-盈利':   '92D050',
    '另类数据-分析师情绪': 'FF78AE',
    '另类数据-微博情绪':  'FF6B6B',
    '另类数据-PMI宏观':   'C9B1FF',
}

def safe(v, default=0):
    try:
        f = float(v)
        return f if not math.isnan(f) and not math.isinf(f) else default
    except:
        return default

# ═══════════════════════════════════════════════════════════════════════════════
# Step 1: 读取三类因子数据
# ═══════════════════════════════════════════════════════════════════════════════

# ── 价量因子 (V7 daily method) ───────────────────────────────────────────────
v7 = pd.read_csv(f'{OUTPUT_DIR}/monthly_ic_pivot_v7.csv')
months = [c for c in v7.columns if c != 'factor']

v7_results = []
for _, row in v7.iterrows():
    factor = row['factor']
    vals = np.array([safe(row[m]) for m in months if pd.notna(row.get(m)) and row.get(m) != ''])
    if len(vals) < 5:
        continue
    ic_mean = safe(vals.mean())
    ic_std  = safe(vals.std())
    ir = ic_mean / ic_std if ic_std != 0 else 0
    pos_rate = (vals > 0).sum() / len(vals)
    abs_ic = abs(ic_mean)
    v7_results.append({
        '因子名称': factor,
        '英文名': factor,
        '大类': '价量技术类',   # 默认，待分类
        '细类': '待分类',
        '平均IC': ic_mean,
        'IC_std': ic_std,
        'IR': ir,
        '|IR|': abs(ir),
        'IC>0月': int((vals > 0).sum()),
        'IC<0月': int((vals < 0).sum()),
        '胜率': pos_rate,
        '绝对IC均': abs_ic,
        '信号方向': '正向↑' if ic_mean > 0 else '逆向↓',
        '月份数': len(vals),
        '来源': 'V7价量因子',
    })

# 补充价量因子分类映射（来自gen_excel_v7.py）
PRICE_CAT = {
    # AlphaForge
    'rank_mom_5': ('AlphaForge AAAI 2025', '动量排名'),
    'rank_mom_10': ('AlphaForge AAAI 2025', '动量排名'),
    'rank_mom_20': ('AlphaGPS KDD 2023', '动量排名'),
    'rank_mom_60': ('AlphaGPS KDD 2023', '动量排名'),
    'rank_prod_20': ('AlphaGen KDD 2023', '量价协同'),
    'rank_prod_60': ('AlphaGen KDD 2023', '量价协同'),
    'rank_prod_af': ('AlphaGen KDD 2023', '量价协同'),
    'rank_ma_ratio_20': ('AlphaForge AAAI 2025', '均线排名'),
    'rank_ma_ratio_60': ('AlphaForge AAAI 2025', '均线排名'),
    'rank_range_5': ('AlphaForge AAAI 2025', '波动率排名'),
    'rank_range_20': ('AlphaForge AAAI 2025', '波动率排名'),
    'rank_range_60': ('AlphaForge AAAI 2025', '波动率排名'),
    'sign_vol20': ('AlphaForge AAAI 2025', '量价信号'),
    'vpt': ('AlphaGen KDD 2023', '量价趋势'),
    'vpt60': ('AlphaGen KDD 2023', '量价趋势'),
    'rsi_ma_10': ('AlphaForge AAAI 2025', 'RSI均值化'),
    # HRFT
    'hrft_sqrt_vol5': ('HRFT WWW 2025', '波动率代理'),
    'hrft_sqrt_vol20': ('HRFT WWW 2025', '波动率代理'),
    'hrft_sqrt_vol60': ('HRFT WWW 2025', '波动率代理'),
    'hrft_abs_log5': ('HRFT WWW 2025', '流动性代理'),
    'hrft_abs_log20': ('HRFT WWW 2025', '流动性代理'),
    'hrft_ewm_vol_short': ('HRFT WWW 2025', '波动率代理'),
    'hrft_ewm_vol_long': ('HRFT WWW 2025', '波动率代理'),
    'hrft_ewm_vol_ratio': ('HRFT WWW 2025', '波动率代理'),
    'hrft_vol_accel': ('HRFT WWW 2025', '波动率动量'),
    'hrft_intraday_atr': ('HRFT WWW 2025', 'ATR代理'),
    # QFR
    'qfr_stable_mom5': ('QFR', '稳定动量'),
    'qfr_stable_mom20': ('QFR', '稳定动量'),
    'qfr_stable_mom60': ('QFR', '稳定动量'),
    'risk_adj_mom5': ('QFR', '风险调整动量'),
    'risk_adj_mom20': ('QFR', '风险调整动量'),
    'risk_adj_mom60': ('QFR', '风险调整动量'),
    'qfr_vol_adj_ret': ('QFR', '波动调整收益'),
    'trend_stability': ('QFR', '趋势稳定'),
    # MASTER
    'master_bias_5_20': ('MASTER', '偏离度'),
    'master_bias_5_60': ('MASTER', '偏离度'),
    'master_bias_20_60': ('MASTER', '偏离度'),
    'master_accel5': ('MASTER', '加速度'),
    'master_accel20': ('MASTER', '加速度'),
    'master_consistency': ('MASTER', '一致性'),
    'master_skew5': ('MASTER', '偏度'),
    'master_skew20': ('MASTER', '偏度'),
    'master_ewm_cross': ('MASTER', '均线交叉'),
    # AlphaSAGE
    'alpha_cov_10': ('AlphaSAGE 2025', '协方差代理'),
    'alpha_cov_20': ('AlphaSAGE 2025', '协方差代理'),
    'alpha_cov_60': ('AlphaSAGE 2025', '协方差代理'),
    'alpha_corr_10': ('AlphaSAGE 2025', '相关性代理'),
    'alpha_corr_20': ('AlphaSAGE 2025', '相关性代理'),
    'alpha_corr_60': ('AlphaSAGE 2025', '相关性代理'),
    # 清华IGSM
    'close52w': ('清华IGSM/北大光华', '52周价位'),
    'close52l': ('清华IGSM/北大光华', '52周价位'),
    'high52w_ratio': ('清华IGSM/北大光华', '52周价位'),
    'lower_shadow': ('清华IGSM/北大光华', 'K线形态'),
    'upper_shadow': ('清华IGSM/北大光华', 'K线形态'),
    'intraday_ret': ('清华IGSM/北大光华', '日内收益'),
    'overnight_ret': ('北大光华', '隔夜收益'),
    'turnover_rate': ('北大光华', '换手率'),
    'turnover_accel': ('北大光华', '换手变化'),
    'amihud': ('北大光华', '非流动性'),
    'intra_vol': ('经典技术因子', '日内波幅'),
    'atr_14': ('经典技术因子', 'ATR'),
    'atr_pos': ('经典技术因子', 'ATR位置'),
    'vol_mv5': ('经典技术因子', '波动率'),
    'vol_mv20': ('经典技术因子', '波动率'),
    'vol_mv60': ('经典技术因子', '波动率'),
    'ma_ratio_5': ('AlphaGPS(KDD2023)', '均线偏离'),
    'ma_ratio_10': ('AlphaGPS(KDD2023)', '均线偏离'),
    'ma_ratio_20': ('AlphaGPS(KDD2023)', '均线偏离'),
    'ma_ratio_60': ('AlphaGPS(KDD2023)', '均线偏离'),
    'ma_ratio_120': ('AlphaGPS(KDD2023)', '均线偏离'),
    'price_pos_20': ('经典技术因子', '价格位置'),
    'price_pos_60': ('经典技术因子', '价格位置'),
    'breakout_20': ('经典技术因子', '突破信号'),
    'breakout_60': ('经典技术因子', '突破信号'),
    'drawdown_20': ('经典技术因子', '回撤'),
    'drawdown_60': ('经典技术因子', '回撤'),
    'drawdown_120': ('经典技术因子', '回撤'),
    'macd': ('经典技术因子', 'MACD'),
    'macd_signal': ('经典技术因子', 'MACD'),
    'macd_hist': ('经典技术因子', 'MACD'),
    'rsi_14': ('经典技术因子', 'RSI'),
    'kdj_k': ('经典技术因子', 'KDJ'),
    'adx_14': ('经典技术因子', 'ADX'),
    '威廉_14': ('经典技术因子', '威廉指标'),
    'bias_5': ('经典技术因子', '乖离率'),
    'bias_20': ('经典技术因子', '乖离率'),
    'bias_60': ('经典技术因子', '乖离率'),
    'bb_pos': ('经典技术因子', '布林带'),
    'kc_pos': ('经典技术因子', '肯特纳'),
    'ma_cross': ('经典技术因子', '均线交叉'),
    'ma_cross_60': ('经典技术因子', '均线交叉'),
    'obv_ma5': ('经典技术因子', 'OBV均线'),
    'obv_ma20': ('经典技术因子', 'OBV均线'),
    'dmi_trend': ('经典技术因子', 'DMI趋势'),
    'aroon_up': ('经典技术因子', 'Aroon'),
    'aroon_down': ('经典技术因子', 'Aroon'),
    'aroon_osc': ('经典技术因子', 'Aroon'),
    'trix_14': ('经典技术因子', 'TRIX'),
    'close_open_gap': ('经典技术因子', '开盘缺口'),
    'range_pct': ('经典技术因子', '振幅'),
    'candle_body': ('经典技术因子', 'K线实体'),
    'ease_of_move': ('AlphaGPS(KDD2023)', '易难动量'),
    'ease_of_move_ma5': ('AlphaGPS(KDD2023)', '易难动量'),
    'fi_ma13': ('经典技术因子', '力指数'),
    'mf_ratio_5': ('经典技术因子', '资金流'),
    'mom_rev': ('经典技术因子', '动量反转'),
    'amt_stability': ('V7新增', '额比稳定'),
    'high_vol_low_vol': ('V7新增', '高波低波'),
    'amount_vol_ratio': ('V7新增', '额量比'),
}

# 补充分类
for r in v7_results:
    fname = r['因子名称']
    if fname in PRICE_CAT:
        r['大类'] = PRICE_CAT[fname][0]
        r['细类'] = PRICE_CAT[fname][1]

v7_df = pd.DataFrame(v7_results)
print(f"价量因子: {len(v7_df)} 个")

# ── 基本面因子 (19个) ─────────────────────────────────────────────────────────
fund_stats = pd.read_csv(f'{OUTPUT_DIR}/fundamental_ic_stats.csv')

# 基本面因子中文名映射
FUND_NAMES = {
    'growth_net_profit': ('基本面-成长', '净利润增长', 'growth_net_profit'),
    'profit_eps': ('基本面-盈利', 'EPS', 'profit_eps'),
    'growth_revenue': ('基本面-成长', '营收增长', 'growth_revenue'),
    'growth_roe': ('基本面-成长', 'ROE增长', 'growth_roe'),
    'quality_roe': ('基本面-质量', 'ROE', 'quality_roe'),
    'quality_roe_diluted': ('基本面-质量', 'ROE稀释', 'quality_roe_diluted'),
    'quality_equity_ratio': ('基本面-质量', '股东权益比', 'quality_equity_ratio'),
    'value_pe': ('基本面-价值', 'PE', 'value_pe'),
    'quality_accumulation': ('基本面-质量', '积累因子', 'quality_accumulation'),
    'value_pb': ('基本面-价值', 'PB', 'value_pb'),
    'quality_undist_profit': ('基本面-质量', '未分配利润', 'quality_undist_profit'),
    'value_pc': ('基本面-价值', 'PC', 'value_pc'),
    'quality_quick_ratio': ('基本面-质量', '速动比率', 'quality_quick_ratio'),
    'value_price': ('基本面-价值', '价格代理', 'value_price'),
    'value_mkt_cap': ('基本面-价值', '市值代理', 'value_mkt_cap'),
    'quality_net_margin': ('基本面-质量', '净利率', 'quality_net_margin'),
    'quality_cps': ('基本面-质量', '现金流', 'quality_cps'),
    'growth_ex_net_profit': ('基本面-成长', '扣非增长', 'growth_ex_net_profit'),
    'profit_bps': ('基本面-盈利', 'BPS', 'profit_bps'),
    'quality_current_ratio': ('基本面-质量', '流动比率', 'quality_current_ratio'),
    'quality_debt_ratio': ('基本面-质量', '资产负债率', 'quality_debt_ratio'),
}

fund_results = []
for _, row in fund_stats.iterrows():
    fname = row['因子']
    if fname not in FUND_NAMES:
        continue
    cat, subcat, ename = FUND_NAMES[fname]
    ic_mean = safe(row['平均IC'])
    ir = safe(row['IR'])
    pos_rate = safe(row['IC>0比例'])
    fund_results.append({
        '因子名称': ename,
        '英文名': ename,
        '大类': cat,
        '细类': subcat,
        '平均IC': ic_mean,
        'IC_std': safe(row['IC_std']),
        'IR': ir,
        '|IR|': abs(ir),
        'IC>0月': int(safe(row['IC>0比例']) * safe(row['月份数'])),
        'IC<0月': int((1 - safe(row['IC>0比例'])) * safe(row['月份数'])),
        '胜率': pos_rate,
        '绝对IC均': safe(row['|IC|']),
        '信号方向': '正向↑' if ic_mean > 0 else '逆向↓',
        '月份数': int(safe(row['月份数'])),
        '来源': 'AKShare财报',
    })

fund_df = pd.DataFrame(fund_results)
print(f"基本面因子: {len(fund_df)} 个")

# ── 另类数据因子 (5个) ────────────────────────────────────────────────────────
alt_ic = pd.read_csv(f'{OUTPUT_DIR}/alternative_ic.csv')

alt_results = []
for _, row in alt_ic.iterrows():
    fname = row['因子名称']
    ic_mean = safe(row['平均IC'])
    ir = safe(row['IR'])
    pos_rate = safe(row['IC>0比例'])
    cat = row.get('大类', '另类数据-分析师情绪')
    alt_results.append({
        '因子名称': fname,
        '英文名': fname,
        '大类': cat,
        '细类': row.get('细类', '分析师'),
        '平均IC': ic_mean,
        'IC_std': safe(row.get('IC_std', 0.1)),
        'IR': ir,
        '|IR|': abs(ir),
        'IC>0月': int(pos_rate * safe(row.get('月份数', 27))),
        'IC<0月': int((1-pos_rate) * safe(row.get('月份数', 27))),
        '胜率': pos_rate,
        '绝对IC均': safe(row.get('|IC|', abs(ic_mean))),
        '信号方向': '正向↑' if ic_mean > 0 else '逆向↓',
        '月份数': int(safe(row.get('月份数', 27))),
        '来源': 'AKShare分析师',
    })

alt_df = pd.DataFrame(alt_results)
print(f"另类因子: {len(alt_df)} 个")

# ═══════════════════════════════════════════════════════════════════════════════
# Step 2: 合并并排名
# ═══════════════════════════════════════════════════════════════════════════════
all_df = pd.concat([v7_df, fund_df, alt_df], ignore_index=True)
all_df = all_df[all_df['月份数'] >= 5]  # 至少5个月IC
all_df = all_df.sort_values('|IR|', ascending=False).reset_index(drop=True)
all_df.insert(0, '排名', range(1, len(all_df)+1))

print(f"\n合并后总因子数: {len(all_df)} 个")

# ═══════════════════════════════════════════════════════════════════════════════
# Step 3: 生成Excel（参考V6格式）
# ═══════════════════════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()
wb.remove(wb.active)

def style_header(cell, text, bg=C_HEADER_BG, fg=C_HEADER_FG, bold=True, size=11):
    cell.value = text
    cell.font = Font(color=fg, bold=bold, size=size)
    cell.fill = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def thin_border():
    thin = Side(style='thin', color='CCCCCC')
    return Border(left=thin, right=thin, top=thin, bottom=thin)

# ── Sheet 1: 因子监控总表 ────────────────────────────────────────────────────
ws1 = wb.create_sheet('因子监控总表')

# 标题行
headers = ['排名', '因子名称', '英文名', '大类', '细类', '平均IC',
           'IC标准差', 'IR', '|IR|', 'IC>0月', 'IC<0月',
           '胜率', '绝对IC均', '信号方向', '月份数', '来源']
for c, h in enumerate(headers, 1):
    style_header(ws1.cell(1, c), h)

ws1.row_dimensions[1].height = 35
ws1.row_dimensions[2].height = 25

# 大类说明行
cat_info_row = 2
for c, h in enumerate(headers, 1):
    ws1.cell(cat_info_row, c).value = ''
ws1.cell(2, 4).value = '← 拖动列筛选'
ws1.cell(2, 4).font = Font(color='808080', italic=True, size=9)

# 数据行
for i, row in all_df.iterrows():
    r = i + 3
    bg = C_ALT_ROW if i % 2 == 0 else C_WHITE

    # 大类颜色
    cat = str(row['大类'])
    cat_color = CAT_COLORS.get(cat, 'FFFFFF')
    cat_fill = PatternFill('solid', fgColor=cat_color)
    cat_font_color = 'FFFFFF' if cat in ['基本面-价值', '清华IGSM'] else '000000'

    vals = [
        row['排名'],
        row['因子名称'],
        row['英文名'],
        row['大类'],
        row['细类'],
        row['平均IC'],
        row['IC_std'],
        row['IR'],
        row['|IR|'],
        row['IC>0月'],
        row['IC<0月'],
        row['胜率'],
        row['绝对IC均'],
        row['信号方向'],
        row['月份数'],
        row['来源'],
    ]

    for c, v in enumerate(vals, 1):
        cell = ws1.cell(r, c)
        cell.value = v
        cell.border = thin_border()
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(size=10)

        if c == 4:  # 大类列用颜色
            cell.fill = cat_fill
            cell.font = Font(color=cat_font_color, size=10, bold=True)
        elif c in [6, 7, 8, 9]:  # IC/IR列
            cell.number_format = '0.0000'
            if isinstance(v, float):
                cell.fill = PatternFill('solid', fgColor='E2EFDA' if v > 0 else 'FCE4D6')
        elif c == 12:  # 胜率
            cell.number_format = '0.0%'
            if isinstance(v, float):
                cell.fill = PatternFill('solid', fgColor='C6EFCE' if v >= 0.5 else 'FFC7CE')
        elif c == 14:  # 信号方向
            cell.font = Font(size=10, bold=True,
                           color='FF0000' if '↑' in str(v) else '00B050')
        else:
            cell.fill = PatternFill('solid', fgColor=bg)

    ws1.row_dimensions[r].height = 18

# 列宽
widths = [6, 22, 18, 18, 12, 10, 10, 8, 8, 8, 8, 8, 10, 10, 8, 14]
for c, w in enumerate(widths, 1):
    set_col_width(ws1, c, w)

# 冻结
ws1.freeze_panes = 'C3'

print("Sheet 1 完成")

# ── Sheet 2: 月度IC热力图（Top30因子 × 40月）─────────────────────────────────
ws2 = wb.create_sheet('月度IC热力图')

# 合并三类数据取月度IC
monthly_ic = {}

# V7月度IC
for _, row in v7.iterrows():
    fname = row['factor']
    for m in months:
        if pd.notna(row.get(m)) and row.get(m) != '':
            if fname not in monthly_ic:
                monthly_ic[fname] = {}
            monthly_ic[fname][m] = safe(row[m])

# 取Top40因子（按|IR|）
top_factors = all_df['因子名称'].tolist()[:40]
top_months = sorted([m for m in months if m >= '2023-01'])[:40]

# 写入
ws2.cell(1, 1).value = '因子/月份'
style_header(ws2.cell(1, 1), '因子/月份')
for c, m in enumerate(top_months, 2):
    cell = ws2.cell(1, c)
    cell.value = m
    style_header(cell, m)
    set_col_width(ws2, c, 8)

set_col_width(ws2, 1, 20)

for r, fname in enumerate(top_factors, 2):
    ws2.cell(r, 1).value = fname
    ws2.cell(r, 1).font = Font(size=9)
    bg = C_ALT_ROW if r % 2 == 0 else C_WHITE
    ws2.cell(r, 1).fill = PatternFill('solid', fgColor=bg)

    for c, m in enumerate(top_months, 2):
        ic_val = monthly_ic.get(fname, {}).get(m, None)
        cell = ws2.cell(r, c)
        if ic_val is not None:
            cell.value = ic_val
            cell.number_format = '0.000'
            # 热力图颜色
            abs_ic = abs(ic_val)
            if abs_ic < 0.01:
                cell.fill = PatternFill('solid', fgColor='F2F2F2')
            elif abs_ic < 0.03:
                cell.fill = PatternFill('solid', fgColor='DDEEFF') if ic_val > 0 else PatternFill('solid', fgColor='FFE0D0')
            elif abs_ic < 0.05:
                cell.fill = PatternFill('solid', fgColor='99CCFF') if ic_val > 0 else PatternFill('solid', fgColor='FF9966')
            else:
                cell.fill = PatternFill('solid', fgColor='0066CC') if ic_val > 0 else PatternFill('solid', fgColor='CC3300')
            cell.font = Font(color='FFFFFF' if abs_ic > 0.05 else '000000', size=8)
        else:
            cell.value = ''
            cell.fill = PatternFill('solid', fgColor='F5F5F5')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border()
    ws2.row_dimensions[r].height = 16

ws2.freeze_panes = 'B2'
print("Sheet 2 完成")

# ── Sheet 3: 使用说明 ───────────────────────────────────────────────────────
ws3 = wb.create_sheet('使用说明')
ws3.column_dimensions['A'].width = 80

instructions = [
    ('★ A股因子IC/IR监控体系 — 全量因子版 (135因子)', True, 14, C_HEADER_BG, 'FFFFFF'),
    ('', False, 10, None, None),
    ('■ 因子库构成', True, 12, '2E75B6', 'FFFFFF'),
    ('  • 价量技术因子: 111个 (技术指标 + AlphaForge + AlphaGPS + AlphaGen + HRFT + QFR + MASTER + AlphaSAGE + 清华IGSM + 北大光华)', False, 10, None, None),
    ('  • 基本面因子: 19个 (AKShare财报: ROE、EPS、营收增长、PE/PB、PC等价值/成长/质量/盈利因子)', False, 10, None, None),
    ('  • 另类数据因子: 5个 (AKShare分析师: 买入比例、关注度、预测EPS、综合评级、EPS预测上调)', False, 10, None, None),
    ('  • 合计: 135个因子', False, 10, None, None),
    ('', False, 10, None, None),
    ('■ IC/IR 指标说明', True, 12, '2E75B6', 'FFFFFF'),
    ('  • IC (Information Coefficient): 因子值与下期收益的相关系数，衡量预测能力', False, 10, None, None),
    ('  • IR (Information Ratio): IC均值/IC标准差，衡量因子稳定性（|IR|>0.5 为有效门槛）', False, 10, None, None),
    ('  • 胜率: IC>0的月份占比，>55%为稳定', False, 10, None, None),
    ('  • 信号方向: 正向↑=因子值越大越强，逆向↓=因子值越小越强', False, 10, None, None),
    ('', False, 10, None, None),
    ('■ 更新频率', True, 12, '2E75B6', 'FFFFFF'),
    ('  • IC统计: 每月末跑一次（价量因子用日线聚合，基本面/另类因子用季报/年报）', False, 10, None, None),
    ('  • 选股: 每日15:30跑一次（用现有因子权重打分）', False, 10, None, None),
    ('  • 因子淘汰: 连续3个月IC变号 或 |IR|<0.3 持续6个月 → 降低权重或淘汰', False, 10, None, None),
    ('', False, 10, None, None),
    ('■ 大类颜色编码', True, 12, '2E75B6', 'FFFFFF'),
    ('  • 蓝: 价量技术类 | 绿: AlphaForge/AlphaGen | 红: HRFT | 橙: AlphaGPS/QFR', False, 10, None, None),
    ('  • 紫: 基本面-价值 | 青: 基本面-成长/质量 | 粉: 另类数据-分析师情绪', False, 10, None, None),
    ('  • 金: 清华IGSM/北大光华', False, 10, None, None),
    ('', False, 10, None, None),
    ('■ 免责声明', True, 12, '2E75B6', 'FFFFFF'),
    ('  • IC/IR基于历史数据，不预示未来表现', False, 10, None, None),
    ('  • 基本面因子季频更新，存在滞后', False, 10, None, None),
    ('  • 选股结果仅供参考，不构成投资建议', False, 10, None, None),
]

for r, (text, bold, size, bg, fg) in enumerate(instructions, 1):
    cell = ws3.cell(r, 1)
    cell.value = text
    if bg:
        cell.fill = PatternFill('solid', fgColor=bg)
        cell.font = Font(color=fg if fg else '000000', bold=bold, size=size)
    else:
        cell.font = Font(size=size, bold=bold)
    cell.alignment = Alignment(horizontal='left', vertical='center')
    ws3.row_dimensions[r].height = 20 if text else 8

print("Sheet 3 完成")

# ── Sheet 4: 通达信选股公式 ─────────────────────────────────────────────────
ws4 = wb.create_sheet('通达信选股公式')
ws4.column_dimensions['A'].width = 15
ws4.column_dimensions['B'].width = 60

formulas = [
    ('★ 通达信选股公式参考', True, 14),
    ('基于135因子IC/IR综合评分，选出各维度最优组合', False, 10),
    ('', False, 10),
    ('【公式A: 价量+基本面综合版（推荐）】', True, 11),
    ('使用因子: overnight_ret(隔夜收益) + turnover_rate(换手率) + rank_prod_60(量价协同) + analyst_eps_2025(预测EPS)', False, 10),
    ('严格版:', True, 10),
    ('XG: (CLOSE/REF(CLOSE,20)-1)>0 AND TURNOVER<5 AND RANK_PROD_60>0 AND ANALYST_EPS_2025>0.5;', False, 9),
    ('宽松版:', True, 10),
    ('XG: CLOSE>MA(CLOSE,5) AND TURNOVER<8;', False, 9),
    ('', False, 10),
    ('【公式B: 基本面优先版】', True, 11),
    ('使用因子: growth_net_profit(净利润增长) + quality_roe(ROE) + profit_eps(EPS)', False, 10),
    ('XG: GROWTH_NET_PROFIT>0.1 AND QUALITY_ROE>10 AND PROFIT_EPS>0.2;', False, 9),
    ('', False, 10),
    ('【公式C: 分析师情绪版】', True, 11),
    ('使用因子: analyst_attention(分析师关注度) + analyst_buy_ratio(买入比例)', False, 10),
    ('XG: ANALYST_ATTENTION>2 AND ANALYST_BUY_RATIO>0.6;', False, 9),
    ('', False, 10),
    ('【公式D: AlphaForge量价协同版】', True, 11),
    ('使用因子: rank_prod_20 + sign_vol20 + vpt', False, 10),
    ('XG: RANK_PROD_20>70 AND SIGN_VOL20<0 AND VPT>0;', False, 9),
    ('', False, 10),
    ('注意: 以上公式需要先在通达信公式管理器中定义各因子变量', False, 9),
]

for r, (text, bold, size) in enumerate(formulas, 1):
    cell = ws4.cell(r, 1)
    cell.value = text
    cell.font = Font(size=size, bold=bold,
                    color=C_HEADER_BG if bold else '000000')
    cell.alignment = Alignment(horizontal='left', vertical='center')
    ws4.row_dimensions[r].height = 18

print("Sheet 4 完成")

# ═══════════════════════════════════════════════════════════════════════════════
# 保存
# ═══════════════════════════════════════════════════════════════════════════════
out_path = f'{OUTPUT_DIR}/因子监控_IC_IR_实测版_全量135因子.xlsx'
wb.save(out_path)
print(f"\n✅ 已保存: {out_path}")

# 打印Top20
print("\n全量因子 Top20 (按|IR|):")
print(all_df[['排名','因子名称','大类','平均IC','IR','胜率','月份数']].head(20).to_string(index=False))
