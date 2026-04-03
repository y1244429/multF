"""
生成因子监控Excel + 可视化图表 + calculate_ic.py模板
基于真实计算的IC/IR数据
"""

import pandas as pd
import numpy as np
import os
import ast
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib import rcParams
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side,
    numbers
)
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter

# ─── 中文字体 ───────────────────────────────────
rcParams['font.sans-serif'] = ['Arial Unicode MS', 'PingFang SC', 'Hiragino Sans GB', 'STHeiti', 'sans-serif']
rcParams['axes.unicode_minus'] = False

OUTPUT_DIR = '/Users/ydy/WorkBuddy/20260326134350/factor_hunter/output'
IC_CSV   = os.path.join(OUTPUT_DIR, 'monthly_ic_matrix.csv')
SUM_CSV  = os.path.join(OUTPUT_DIR, 'factor_ic_summary.csv')

# ─── 加载数据 ─────────────────────────────────────
ic_df   = pd.read_csv(IC_CSV, index_col=0)
summary = pd.read_csv(SUM_CSV)

# 补充状态列（如缺失）
def get_status(row):
    ic = row.get('近3月均IC', np.nan)
    ir = row.get('近3月IR',  np.nan)
    if pd.isna(ic) or pd.isna(ir): return '⚪数据不足'
    if ic > 0.1 and ir >= 1.0:    return '✅健康'
    elif ic > 0.05 and ir >= 0.5: return '⚠️一般'
    elif ic < 0 or ir < 0.3:      return '❌停用'
    else:                          return '🟡预警'

if '状态' not in summary.columns:
    summary['状态'] = summary.apply(get_status, axis=1)

# 解析近6月IC历史（字符串→list）
def parse_list(s):
    try:
        v = ast.literal_eval(str(s))
        return [float(x) if not (isinstance(x, float) and np.isnan(x)) else np.nan for x in v]
    except:
        return []

summary['近6月IC历史_list'] = summary['近6月IC历史'].apply(parse_list)

months     = ic_df.columns.tolist()
recent6    = months[-6:]
recent3    = months[-3:]
all_months = months  # 全部历史月份（用于原始数据表）

# ─── 颜色 ──────────────────────────────────────
C_HEADER   = 'FF1F3D6B'   # 深蓝
C_SUBHDR   = 'FF2E5FAA'   # 蓝
C_TOP5     = 'FFFFEEBA'   # 金黄
C_GREEN    = 'FFD6F5D6'   # 健康绿
C_YELLOW   = 'FFFFF3CD'   # 一般黄
C_ORANGE   = 'FFFFE5CC'   # 预警橙
C_RED      = 'FFFFDADA'   # 停用红
C_GRAY     = 'FFF5F5F5'   # 交替行背景
C_WHITE    = 'FFFFFFFF'

def fill(hex6):
    return PatternFill('solid', fgColor=hex6)

def bold_font(color='FF000000', size=11):
    return Font(bold=True, color=color, size=size)

def center():
    return Alignment(horizontal='center', vertical='center', wrap_text=True)

thin = Side(style='thin', color='FFB0B0B0')
def border():
    return Border(left=thin, right=thin, top=thin, bottom=thin)

# ─── 状态→颜色映射 ────────────────────────────────
def status_fill(status):
    if '健康' in str(status): return fill(C_GREEN)
    if '一般' in str(status): return fill(C_YELLOW)
    if '预警' in str(status): return fill(C_ORANGE)
    if '停用' in str(status): return fill(C_RED)
    return fill(C_WHITE)

# ─── 建 Workbook ──────────────────────────────────
wb = openpyxl.Workbook()

# ══════════════════════════════════════════════════
# SHEET 1：因子监控（主表）
# ══════════════════════════════════════════════════
ws1 = wb.active
ws1.title = '因子监控'

# -- 标题行
ws1.merge_cells('A1:J1')
ws1['A1'] = 'A股因子监控表（IC/IR 实时追踪）'
ws1['A1'].font      = bold_font('FFFFFFFF', 14)
ws1['A1'].fill      = fill(C_HEADER)
ws1['A1'].alignment = center()
ws1.row_dimensions[1].height = 36

# -- 副标题
ws1.merge_cells('A2:J2')
ws1['A2'] = f'数据基础：market_data_v2，600只A股，{ic_df.columns[0]}～{ic_df.columns[-1]}，共{len(ic_df.columns)}期'
ws1['A2'].font      = Font(size=10, color='FF555555', italic=True)
ws1['A2'].alignment = center()
ws1.row_dimensions[2].height = 20

# -- 列头
headers = ['排名', '因子', '类别', '本月IC', '近3月均IC', '近3月IR', '近6月IR', '状态', '建议权重', '近6月IC走势（→最新）']
for col, h in enumerate(headers, 1):
    cell = ws1.cell(row=3, column=col, value=h)
    cell.font      = bold_font('FFFFFFFF', 10)
    cell.fill      = fill(C_SUBHDR)
    cell.alignment = center()
    cell.border    = border()
ws1.row_dimensions[3].height = 28

# 列宽
col_widths = [6, 16, 12, 10, 12, 10, 10, 12, 12, 45]
for i, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# 因子类别映射
CATEGORY = {
    'mom_5':'动量', 'mom_10':'动量', 'mom_20':'动量', 'mom_60':'动量',
    'ma_ratio_5':'均线', 'ma_ratio_10':'均线', 'ma_ratio_20':'均线', 'ma_ratio_60':'均线',
    'vol_10':'波动率', 'vol_20':'波动率', 'vol_60':'波动率',
    'v_ratio':'成交量', 'v_mom_5':'成交量', 'v_mom_20':'成交量',
    'amt_stability':'成交量', 'low_turnover':'成交量',
    'price_pos_10':'价格位置', 'price_pos_20':'价格位置', 'price_pos_60':'价格位置',
    'breakout_20':'突破', 'breakout_60':'突破',
    'drawdown_20':'回撤', 'drawdown_60':'回撤',
    'candle_body':'K线形态', 'upper_shadow':'K线形态', 'lower_shadow':'K线形态',
}

# 建议权重逻辑
def suggest_weight(rank, status):
    if '停用' in str(status): return '0%'
    if rank <= 3: return '25%'
    if rank <= 5: return '20%'
    if rank <= 10: return '10%'
    return '5%'

# 数据行
for i, row in summary.iterrows():
    r = i + 4
    rank    = int(row['排名'])
    factor  = row['因子']
    status  = row['状态']
    ic6_list = row['近6月IC历史_list']
    bg      = fill(C_TOP5) if rank <= 5 else (fill(C_GRAY) if i % 2 == 0 else fill(C_WHITE))

    values = [
        rank,
        factor,
        CATEGORY.get(factor, '其他'),
        row['本月IC'],
        row['近3月均IC'],
        row['近3月IR'],
        row['近6月IR'],
        status,
        suggest_weight(rank, status),
        '  '.join([f'{v:+.3f}' if not np.isnan(v) else 'N/A' for v in ic6_list]),
    ]
    for col, val in enumerate(values, 1):
        cell = ws1.cell(row=r, column=col, value=val)
        cell.alignment = center()
        cell.border    = border()
        # Top5金色背景，否则交替灰白
        if rank <= 5:
            cell.fill = fill(C_TOP5)
        else:
            cell.fill = bg
        # 数值格式
        if col in (4, 5):
            cell.number_format = '0.000'
        if col in (6, 7):
            cell.number_format = '0.00'
        # 状态列着色
        if col == 8:
            cell.fill = status_fill(status)
            cell.font = bold_font(size=10)

ws1.row_dimensions[r].height = 22

# -- 判断规则说明
last_data_row = 3 + len(summary)
ws1.merge_cells(f'A{last_data_row+2}:J{last_data_row+2}')
note_cell = ws1.cell(row=last_data_row+2, column=1,
    value='【判断规则】✅健康：近3月均IC>0.10 且 IR≥1.0  |  ⚠️一般：IC>0.05 且 IR≥0.5  |  🟡预警：其他  |  ❌停用：IC<0 或 IR<0.3')
note_cell.font = Font(size=9, color='FF555555', italic=True)
note_cell.alignment = Alignment(horizontal='left', vertical='center')
ws1.row_dimensions[last_data_row+2].height = 20

# ══════════════════════════════════════════════════
# SHEET 2：Top5详情
# ══════════════════════════════════════════════════
ws2 = wb.create_sheet('Top5因子详情')
ws2.merge_cells('A1:H1')
ws2['A1'] = 'IC/IR 排名 Top5 因子详细数据'
ws2['A1'].font = bold_font('FFFFFFFF', 13)
ws2['A1'].fill = fill(C_HEADER)
ws2['A1'].alignment = center()
ws2.row_dimensions[1].height = 32

top5 = summary.head(5).reset_index(drop=True)

# 各Top5因子的月度IC走势
for idx, (_, row) in enumerate(top5.iterrows()):
    start_col = idx * 9 + 1
    factor = row['因子']

    # 因子标题
    ws2.merge_cells(start_row=3, start_column=start_col, end_row=3, end_column=start_col+7)
    cell = ws2.cell(row=3, column=start_col,
        value=f"#{int(row['排名'])} {factor}  [{CATEGORY.get(factor,'其他')}]  {row['状态']}")
    cell.font = bold_font('FFFFFFFF', 11)
    cell.fill = fill(C_SUBHDR)
    cell.alignment = center()

    # 统计摘要
    stats_rows = [
        ('本月IC', f"{row['本月IC']:.4f}"),
        ('近3月均IC', f"{row['近3月均IC']:.4f}"),
        ('近3月IR', f"{row['近3月IR']:.2f}"),
        ('近6月IR', f"{row['近6月IR']:.2f}"),
        ('状态', row['状态']),
        ('建议权重', suggest_weight(int(row['排名']), row['状态'])),
    ]
    for sr, (k, v) in enumerate(stats_rows, 4):
        ws2.cell(row=sr, column=start_col, value=k).font = bold_font(size=10)
        ws2.cell(row=sr, column=start_col, value=k).fill = fill(C_GRAY)
        ws2.cell(row=sr, column=start_col+1, value=v).alignment = center()
        ws2.cell(row=sr, column=start_col+1, value=v).fill = status_fill(row['状态']) if k == '状态' else fill(C_WHITE)

    # 月度IC历史（最近24个月）
    ws2.cell(row=11, column=start_col, value='月份').font = bold_font(size=9)
    ws2.cell(row=11, column=start_col+1, value='IC').font = bold_font(size=9)
    recent24 = all_months[-24:]
    for mr, m in enumerate(recent24, 12):
        ic_val = ic_df.loc[factor, m] if m in ic_df.columns else np.nan
        ws2.cell(row=mr, column=start_col, value=m).font = Font(size=9)
        ic_cell = ws2.cell(row=mr, column=start_col+1, value=round(ic_val, 4) if not np.isnan(ic_val) else '')
        ic_cell.number_format = '0.000'
        if not np.isnan(ic_val):
            if ic_val > 0.1: ic_cell.fill = fill(C_GREEN)
            elif ic_val > 0.05: ic_cell.fill = fill(C_YELLOW)
            elif ic_val < 0: ic_cell.fill = fill(C_RED)

    for ci in range(start_col, start_col+8):
        ws2.column_dimensions[get_column_letter(ci)].width = 11

# ══════════════════════════════════════════════════
# SHEET 3：原始数据（全部月度IC）
# ══════════════════════════════════════════════════
ws3 = wb.create_sheet('原始数据（月度IC）')
ws3.merge_cells('A1:B1')
ws3['A1'] = 'IC月度原始数据（用于手动更新或校验）'
ws3['A1'].font = bold_font('FFFFFFFF', 12)
ws3['A1'].fill = fill(C_HEADER)
ws3['A1'].alignment = center()

# 表头：因子名 + 月份列
ws3.cell(row=2, column=1, value='因子').font = bold_font(size=10)
ws3.cell(row=2, column=1).fill = fill(C_SUBHDR)
for ci, m in enumerate(all_months, 2):
    c = ws3.cell(row=2, column=ci, value=m)
    c.font = bold_font('FFFFFFFF', 9)
    c.fill = fill(C_SUBHDR)
    c.alignment = center()
    ws3.column_dimensions[get_column_letter(ci)].width = 10

ws3.column_dimensions['A'].width = 18

for ri, factor in enumerate(ic_df.index, 3):
    ws3.cell(row=ri, column=1, value=factor)
    bg = fill(C_GRAY) if ri % 2 == 0 else fill(C_WHITE)
    ws3.cell(row=ri, column=1).fill = bg
    for ci, m in enumerate(all_months, 2):
        val = ic_df.loc[factor, m] if m in ic_df.columns else np.nan
        cell = ws3.cell(row=ri, column=ci, value=round(val, 4) if not np.isnan(val) else '')
        cell.number_format = '0.000'
        cell.alignment = center()
        if not np.isnan(val):
            if   val >  0.10: cell.fill = fill(C_GREEN)
            elif val >  0.05: cell.fill = fill('FFDAEEDD')
            elif val < -0.05: cell.fill = fill(C_RED)
            elif val <  0:    cell.fill = fill('FFFFEEEE')
            else:             cell.fill = bg

# ══════════════════════════════════════════════════
# SHEET 4：使用说明
# ══════════════════════════════════════════════════
ws4 = wb.create_sheet('使用说明')
ws4.column_dimensions['A'].width = 80

lines = [
    ('A股因子监控模板 使用手册', True, 14, C_HEADER, 'FFFFFFFF'),
    ('', False, 10, C_WHITE, 'FF000000'),
    ('一、什么是IC/IR？', True, 12, C_SUBHDR, 'FFFFFFFF'),
    ('IC（Information Coefficient）= 因子截面排名 与 未来20日收益排名 的 Spearman相关系数', False, 10, C_WHITE, 'FF000000'),
    ('  • 范围：-1 到 +1，越大说明因子预测能力越强', False, 10, C_WHITE, 'FF000000'),
    ('  • IC > 0.15：很强   IC 0.05~0.15：有效   IC < 0.05：弱效   IC < 0：反向', False, 10, C_WHITE, 'FF000000'),
    ('IR（Information Ratio）= 近N期IC均值 / 近N期IC标准差', False, 10, C_WHITE, 'FF000000'),
    ('  • IR 衡量 IC 的稳定性，IR越高说明因子预测越稳定', False, 10, C_WHITE, 'FF000000'),
    ('  • IR ≥ 1.0：稳定   IR 0.5~1.0：一般   IR < 0.5：不稳定', False, 10, C_WHITE, 'FF000000'),
    ('', False, 10, C_WHITE, 'FF000000'),
    ('二、判断标准', True, 12, C_SUBHDR, 'FFFFFFFF'),
    ('  ✅ 健康（绿色）：近3月均IC > 0.10  且  IR ≥ 1.0 → 正常使用，按建议权重配置', False, 10, C_GREEN, 'FF006600'),
    ('  ⚠️ 一般（黄色）：近3月均IC > 0.05  且  IR ≥ 0.5 → 减半使用，加强监控', False, 10, C_YELLOW, 'FF665500'),
    ('  🟡 预警（橙色）：IC或IR未达到一般标准 → 降权使用，观察1-2个月', False, 10, C_ORANGE, 'FF884400'),
    ('  ❌ 停用（红色）：IC < 0  或  IR < 0.3 → 立即停用，等待市场风格切换', False, 10, C_RED, 'FF880000'),
    ('', False, 10, C_WHITE, 'FF000000'),
    ('三、每月操作流程（3步）', True, 12, C_SUBHDR, 'FFFFFFFF'),
    ('  第1步：运行 calc_ic_realdata.py（约5分钟），获得最新月度IC矩阵', False, 10, C_WHITE, 'FF000000'),
    ('  第2步：将新一列IC数据粘贴到"原始数据"表，保存', False, 10, C_WHITE, 'FF000000'),
    ('  第3步：查看"因子监控"表，找Top5因子，调整通达信选股公式权重', False, 10, C_WHITE, 'FF000000'),
    ('', False, 10, C_WHITE, 'FF000000'),
    ('四、当前Top5因子（实测数据）', True, 12, C_SUBHDR, 'FFFFFFFF'),
]
for r, (txt, is_bold, fsize, bg_c, fg_c) in enumerate(lines, 1):
    cell = ws4.cell(row=r, column=1, value=txt)
    cell.font = Font(bold=is_bold, size=fsize, color=fg_c)
    cell.fill = fill(bg_c)
    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws4.row_dimensions[r].height = 22 if txt else 8

# 追加Top5详情
start_r = len(lines) + 1
top5_headers = ['排名', '因子', '类别', '本月IC', '近3月均IC', '近3月IR', '状态']
for ci, h in enumerate(top5_headers, 1):
    c = ws4.cell(row=start_r, column=ci, value=h)
    c.font = bold_font('FFFFFFFF', 10)
    c.fill = fill(C_SUBHDR)
    c.alignment = center()
    ws4.column_dimensions[get_column_letter(ci)].width = 14
ws4.column_dimensions['A'].width = 6

for i, (_, row) in enumerate(top5.iterrows(), 1):
    r = start_r + i
    for ci, val in enumerate([int(row['排名']), row['因子'], CATEGORY.get(row['因子'],'其他'),
                               row['本月IC'], row['近3月均IC'], row['近3月IR'], row['状态']], 1):
        cell = ws4.cell(row=r, column=ci, value=val)
        cell.alignment = center()
        cell.fill = fill(C_TOP5)
        if ci in (4,5): cell.number_format = '0.000'
        if ci == 6: cell.number_format = '0.00'
        if ci == 7: cell.fill = status_fill(row['状态'])

# ─── 保存 Excel ──────────────────────────────────
XLSX_PATH = os.path.join(OUTPUT_DIR, '因子监控_IC_IR_实测版.xlsx')
wb.save(XLSX_PATH)
print(f"✅ Excel 已保存: {XLSX_PATH}")

# ══════════════════════════════════════════════════
# 可视化：IC/IR监控图（Top5 + 全部因子分布）
# ══════════════════════════════════════════════════
print("生成可视化图表...")
fig = plt.figure(figsize=(20, 24))
fig.patch.set_facecolor('#F8F9FA')

top5_factors = top5['因子'].tolist()
top5_colors  = ['#E63946', '#2A9D8F', '#E9C46A', '#F4A261', '#264653']

# ── 子图1：Top5因子月度IC走势（全部历史）
ax1 = fig.add_subplot(4, 1, 1)
ax1.set_facecolor('#FFFFFF')
ax1.axhline(0.10, color='#2A9D8F', linestyle='--', linewidth=1.5, alpha=0.8, label='健康线(0.10)')
ax1.axhline(0.05, color='#E9C46A', linestyle='--', linewidth=1.5, alpha=0.8, label='警戒线(0.05)')
ax1.axhline(0,    color='#E63946', linestyle='-',  linewidth=1.0, alpha=0.5)
ax1.fill_between(range(len(all_months)), 0.10, 0.30, alpha=0.08, color='#2A9D8F')
ax1.fill_between(range(len(all_months)), 0.05, 0.10, alpha=0.08, color='#E9C46A')

for factor, color in zip(top5_factors, top5_colors):
    ic_vals = [ic_df.loc[factor, m] if m in ic_df.columns else np.nan for m in all_months]
    ax1.plot(range(len(all_months)), ic_vals, color=color, linewidth=2.2,
             marker='o', markersize=4, label=f'{factor}', alpha=0.9)

ax1.set_xticks(range(0, len(all_months), 3))
ax1.set_xticklabels([all_months[i] for i in range(0, len(all_months), 3)], rotation=45, fontsize=9)
ax1.set_title('Top5 因子月度IC走势（全历史）', fontsize=14, fontweight='bold', pad=12)
ax1.set_ylabel('IC值', fontsize=11)
ax1.legend(loc='upper right', fontsize=9, framealpha=0.9)
ax1.grid(True, alpha=0.3)
ax1.set_ylim(-0.4, 0.35)

# ── 子图2：近6月IC对比（所有因子，按近3月均IC排序）
ax2 = fig.add_subplot(4, 1, 2)
ax2.set_facecolor('#FFFFFF')
n_factors = len(summary)
x = np.arange(n_factors)
bar_colors = []
for _, row in summary.iterrows():
    if '健康' in str(row['状态']): bar_colors.append('#2A9D8F')
    elif '一般' in str(row['状态']): bar_colors.append('#E9C46A')
    elif '停用' in str(row['状态']): bar_colors.append('#E63946')
    else: bar_colors.append('#F4A261')

bars = ax2.bar(x, summary['近3月均IC'], color=bar_colors, alpha=0.85, width=0.7)
ax2.axhline(0.10, color='#2A9D8F', linestyle='--', linewidth=1.5, alpha=0.8, label='健康线(0.10)')
ax2.axhline(0.05, color='#E9C46A', linestyle='--', linewidth=1.5, alpha=0.8, label='警戒线(0.05)')
ax2.axhline(0,    color='black',   linestyle='-',  linewidth=0.8, alpha=0.5)

# 标注Top5序号
for i, (_, row) in enumerate(summary.iterrows()):
    if row['排名'] <= 5:
        ax2.text(i, row['近3月均IC'] + 0.008, f"#{int(row['排名'])}", ha='center', va='bottom', fontsize=8, fontweight='bold', color='#1F3D6B')

ax2.set_xticks(x)
ax2.set_xticklabels(summary['因子'].tolist(), rotation=60, ha='right', fontsize=8)
ax2.set_title('所有因子 近3月均IC 对比（按评分排序）', fontsize=14, fontweight='bold', pad=12)
ax2.set_ylabel('近3月均IC', fontsize=11)
legend_patches = [
    mpatches.Patch(color='#2A9D8F', label='✅ 健康'),
    mpatches.Patch(color='#E9C46A', label='⚠️ 一般'),
    mpatches.Patch(color='#F4A261', label='🟡 预警'),
    mpatches.Patch(color='#E63946', label='❌ 停用'),
]
ax2.legend(handles=legend_patches, loc='upper right', fontsize=9)
ax2.grid(True, axis='y', alpha=0.3)

# ── 子图3：近3月IR 条形图
ax3 = fig.add_subplot(4, 1, 3)
ax3.set_facecolor('#FFFFFF')
ir_vals = summary['近3月IR'].fillna(0).tolist()
ir_colors = ['#2A9D8F' if v >= 1.0 else ('#E9C46A' if v >= 0.5 else ('#F4A261' if v >= 0.3 else '#E63946')) for v in ir_vals]
ax3.bar(x, ir_vals, color=ir_colors, alpha=0.85, width=0.7)
ax3.axhline(1.0, color='#2A9D8F', linestyle='--', linewidth=1.5, alpha=0.8, label='健康IR(1.0)')
ax3.axhline(0.5, color='#E9C46A', linestyle='--', linewidth=1.5, alpha=0.8, label='警戒IR(0.5)')
ax3.axhline(0,   color='black',   linestyle='-',  linewidth=0.8, alpha=0.5)
ax3.set_xticks(x)
ax3.set_xticklabels(summary['因子'].tolist(), rotation=60, ha='right', fontsize=8)
ax3.set_title('所有因子 近3月IR（稳定性）对比', fontsize=14, fontweight='bold', pad=12)
ax3.set_ylabel('近3月IR', fontsize=11)
ax3.legend(loc='upper right', fontsize=9)
ax3.grid(True, axis='y', alpha=0.3)

# ── 子图4：IC散点图（近3月均IC vs 近3月IR）
ax4 = fig.add_subplot(4, 1, 4)
ax4.set_facecolor('#FFFFFF')
ax4.axhline(1.0, color='#2A9D8F', linestyle='--', linewidth=1.2, alpha=0.7)
ax4.axhline(0.5, color='#E9C46A', linestyle='--', linewidth=1.2, alpha=0.7)
ax4.axvline(0.10, color='#2A9D8F', linestyle='--', linewidth=1.2, alpha=0.7)
ax4.axvline(0.05, color='#E9C46A', linestyle='--', linewidth=1.2, alpha=0.7)
ax4.axvline(0,    color='#E63946', linestyle='-',  linewidth=0.8, alpha=0.5)

# 四象限填色
ax4.fill_between([0.10, 0.5], [1.0, 1.0], [8, 8], alpha=0.08, color='#2A9D8F')  # 健康区
ax4.fill_between([0.05, 0.10],[0.5, 0.5], [8, 8], alpha=0.08, color='#E9C46A')  # 一般区
ax4.text(0.30, 6.5, '✅ 健康区', fontsize=11, color='#2A9D8F', fontweight='bold', alpha=0.8)
ax4.text(0.06, 5.5, '⚠️ 一般区', fontsize=10, color='#B8860B', fontweight='bold', alpha=0.8)

for _, row in summary.iterrows():
    ic_ = row['近3月均IC']
    ir_ = row['近3月IR']
    if pd.isna(ic_) or pd.isna(ir_): continue
    color = '#2A9D8F' if '健康' in str(row['状态']) else ('#E9C46A' if '一般' in str(row['状态']) else ('#F4A261' if '预警' in str(row['状态']) else '#E63946'))
    size  = 200 if row['排名'] <= 5 else 80
    ax4.scatter(ic_, ir_, c=color, s=size, alpha=0.85, zorder=5)
    if row['排名'] <= 5:
        ax4.annotate(f"#{int(row['排名'])}\n{row['因子']}", (ic_, ir_),
                     textcoords='offset points', xytext=(8, 4),
                     fontsize=8, fontweight='bold', color='#1F3D6B',
                     arrowprops=dict(arrowstyle='->', color='#888888', lw=1))

ax4.set_xlabel('近3月均IC', fontsize=11)
ax4.set_ylabel('近3月IR', fontsize=11)
ax4.set_title('IC vs IR 二维象限图（大点=Top5）', fontsize=14, fontweight='bold', pad=12)
ax4.grid(True, alpha=0.3)

plt.tight_layout(pad=3.0)
VIZ_PATH = os.path.join(OUTPUT_DIR, 'factor_ic_ir_monitor.png')
plt.savefig(VIZ_PATH, dpi=150, bbox_inches='tight', facecolor='#F8F9FA')
plt.close()
print(f"✅ 可视化图表已保存: {VIZ_PATH}")

# ══════════════════════════════════════════════════
# calculate_ic.py 完整模板（对接market_data_v2）
# ══════════════════════════════════════════════════
CALC_TEMPLATE = '''"""
calculate_ic.py  — A股因子IC月度计算器（完整版）
基于 market_data_v2 真实数据，每月运行一次，输出：
  1. 各因子本月IC值
  2. 近3/6月均IC
  3. 近3/6月IR
  4. Top5因子排名
  
使用方法：
  python calculate_ic.py
  python calculate_ic.py --top 5      # 只打印Top5
  python calculate_ic.py --month 2025-11  # 指定月份
"""

import os, glob, argparse
import pandas as pd
import numpy as np
from scipy import stats
from datetime import datetime
import warnings
warnings.filterwarnings('ignore')

DATA_DIR    = '/Users/ydy/Downloads/market_data_v2'
IC_MATRIX   = '/Users/ydy/WorkBuddy/20260326134350/factor_hunter/output/monthly_ic_matrix.csv'
OUTPUT_DIR  = '/Users/ydy/WorkBuddy/20260326134350/factor_hunter/output'

# ─── 核心：单因子IC计算（Spearman秩相关）────────────────────
def calc_ic(factor_series: pd.Series, forward_ret: pd.Series) -> float:
    """
    计算 Spearman IC
    参数:
        factor_series: 截面因子值（当天所有股票）
        forward_ret:   对应股票的未来N日收益
    返回:
        ic: float，[-1, 1]
    """
    df = pd.DataFrame({"f": factor_series, "r": forward_ret}).dropna()
    if len(df) < 30:
        return np.nan
    ic, _ = stats.spearmanr(df["f"], df["r"])
    return round(ic, 4)


# ─── 因子计算函数（逐股票）────────────────────────────────
def compute_stock_factors(df: pd.DataFrame) -> pd.DataFrame:
    """
    输入单只股票的OHLCV数据，输出每日因子值
    df columns: date, open, high, low, close, volume, amount
    """
    c = df["close"].values.astype(float)
    h = df["high"].values.astype(float)
    l = df["low"].values.astype(float)
    v = df["volume"].values.astype(float)
    a = df["amount"].values.astype(float)
    n = len(c)

    def sma(arr, w):
        res = np.full(n, np.nan)
        for i in range(w-1, n):
            res[i] = arr[max(0,i-w+1):i+1].mean()
        return res

    def smax(arr, w):
        res = np.full(n, np.nan)
        for i in range(w-1, n):
            res[i] = arr[max(0,i-w+1):i+1].max()
        return res

    def smin(arr, w):
        res = np.full(n, np.nan)
        for i in range(w-1, n):
            res[i] = arr[max(0,i-w+1):i+1].min()
        return res

    def sstd(arr, w):
        res = np.full(n, np.nan)
        for i in range(w-1, n):
            res[i] = arr[max(0,i-w+1):i+1].std(ddof=1)
        return res

    def pct(arr, p):
        r = np.full(n, np.nan)
        sh = np.roll(arr, p)
        r[p:] = np.where(np.abs(sh[p:])>1e-9, (arr[p:]-sh[p:])/np.abs(sh[p:]), np.nan)
        return r

    res = pd.DataFrame(index=df.index)
    res["date"] = df["date"].values
    res["code"] = df["code"].values if "code" in df.columns else "unknown"

    # 动量
    for p in [5, 10, 20, 60]:
        res[f"mom_{p}"] = pct(c, p)

    # 均线偏离
    for p in [5, 10, 20, 60]:
        ma = sma(c, p)
        res[f"ma_ratio_{p}"] = np.where(np.abs(ma)>1e-9, (c-ma)/np.abs(ma), np.nan)

    # 波动率
    log_ret = np.diff(np.log(np.where(c>0, c, np.nan)), prepend=np.nan)
    for p in [10, 20, 60]:
        res[f"vol_{p}"] = sstd(log_ret, p)

    # 成交量因子
    v_ma20 = sma(v, 20)
    res["v_ratio"] = np.where(np.abs(v_ma20)>1e-9, v/np.abs(v_ma20), np.nan)
    res["low_turnover"] = -res["v_ratio"]  # 低换手
    res["v_mom_5"]  = pct(v, 5)
    res["v_mom_20"] = pct(v, 20)
    a_ma20 = sma(a, 20)
    res["amt_stability"] = np.where(np.abs(a_ma20)>1e-9, a/np.abs(a_ma20), np.nan)

    # 价格位置
    for p in [10, 20, 60]:
        hh = smax(h, p); ll = smin(l, p)
        denom = hh - ll
        res[f"price_pos_{p}"] = np.where(denom>1e-9, (c-ll)/denom, np.nan)

    # 突破信号
    for p in [20, 60]:
        hh = smax(h, p)
        res[f"breakout_{p}"] = np.where(np.abs(hh)>1e-9, (c-hh)/np.abs(hh), np.nan)

    # 回撤
    for p in [20, 60]:
        hh = smax(h, p)
        res[f"drawdown_{p}"] = np.where(np.abs(hh)>1e-9, (c-hh)/np.abs(hh), np.nan)

    # K线形态
    hl = h - l; hl[hl<1e-9] = np.nan
    prev_c = np.roll(c, 1); prev_c[0] = np.nan
    res["candle_body"]  = np.abs(c-prev_c)/hl
    res["upper_shadow"] = (h - np.maximum(c, prev_c))/hl
    res["lower_shadow"] = (np.minimum(c, prev_c) - l)/hl

    return res


# ─── 加载数据 ──────────────────────────────────────
def load_data(n_stocks=600):
    files = sorted(glob.glob(os.path.join(DATA_DIR, "*.csv")))[:n_stocks]
    frames = []
    for f in files:
        try:
            df = pd.read_csv(f, parse_dates=["date"])
            df["code"] = os.path.basename(f).replace(".csv","")
            frames.append(df)
        except:
            pass
    data = pd.concat(frames, ignore_index=True)
    data = data.sort_values(["code","date"]).reset_index(drop=True)
    for col in ["open","high","low","close","volume","amount"]:
        data[col] = data[col].astype(float)
    return data


# ─── 计算指定月份的截面IC ─────────────────────────────
def calc_monthly_ic(target_month: str, forward_days: int = 20, n_stocks: int = 600) -> dict:
    """
    计算 target_month（格式：2025-11）这个月末截面的各因子IC
    返回 {factor_name: ic_value}
    """
    print(f"加载数据（{n_stocks}只股票）...")
    data = load_data(n_stocks)

    print("计算因子...")
    factor_frames = []
    for code, g in data.groupby("code"):
        if len(g) < 80: continue
        try:
            factor_frames.append(compute_stock_factors(g.reset_index(drop=True)))
        except:
            pass
    factor_data = pd.concat(factor_frames, ignore_index=True)
    factor_data["date"] = pd.to_datetime(factor_data["date"])
    FACTOR_COLS = [c for c in factor_data.columns if c not in ["date","code"]]

    # 未来收益
    data_s = data.sort_values(["code","date"])
    data_s["fwd_ret"] = data_s.groupby("code")["close"].transform(
        lambda x: x.shift(-forward_days)/x - 1)
    fwd = data_s[["code","date","fwd_ret"]].copy()
    fwd["date"] = pd.to_datetime(fwd["date"])
    factor_data = factor_data.merge(fwd, on=["code","date"], how="left")

    # 取目标月末截面
    factor_data["ym"] = factor_data["date"].dt.to_period("M")
    month_data = factor_data[factor_data["ym"].astype(str) == target_month]
    if month_data.empty:
        print(f"警告：{target_month} 无数据")
        return {}
    last_date = month_data["date"].max()
    cross = factor_data[factor_data["date"] == last_date].dropna(subset=["fwd_ret"])
    print(f"  截面日期：{last_date.date()}，股票数：{len(cross)}")

    # 计算IC
    results = {}
    for factor in FACTOR_COLS:
        sub = cross[[factor, "fwd_ret"]].dropna()
        if len(sub) < 20:
            results[factor] = np.nan
            continue
        ic, _ = stats.spearmanr(sub[factor], sub["fwd_ret"])
        results[factor] = round(ic, 4)

    return results


# ─── IR计算 ──────────────────────────────────────────
def calc_ir(ic_series: list) -> float:
    ic = [x for x in ic_series if not np.isnan(x)]
    if len(ic) < 2: return np.nan
    return round(np.mean(ic) / (np.std(ic, ddof=0) + 1e-9), 4)


# ─── 主程序 ──────────────────────────────────────────
if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--top",   type=int, default=5, help="打印Top N因子")
    parser.add_argument("--month", type=str, default=None, help="指定月份，格式2025-11，默认最新")
    args = parser.parse_args()

    # 加载历史IC矩阵
    if os.path.exists(IC_MATRIX):
        ic_df = pd.read_csv(IC_MATRIX, index_col=0)
        print(f"历史IC矩阵：{ic_df.shape[0]}个因子，{ic_df.shape[1]}个月")
        all_months = ic_df.columns.tolist()
    else:
        ic_df = pd.DataFrame()
        all_months = []

    # 确定目标月份
    if args.month:
        target_month = args.month
    else:
        # 默认计算下一个未记录月份
        from pandas.tseries.offsets import MonthEnd
        if all_months:
            last_m = pd.Period(all_months[-1], "M")
            target_month = str(last_m + 1)
        else:
            target_month = pd.Timestamp.now().to_period("M").strftime("%Y-%m")

    print(f"\\n===== 计算 {target_month} 截面IC =====")
    ic_results = calc_monthly_ic(target_month)

    if not ic_results:
        print("无法计算，请检查数据")
        exit(1)

    # 更新IC矩阵
    for factor, ic in ic_results.items():
        if len(ic_df) == 0 or factor not in ic_df.index:
            ic_df.loc[factor, target_month] = ic
        else:
            ic_df.loc[factor, target_month] = ic
    ic_df.to_csv(IC_MATRIX)
    print(f"✅ IC矩阵已更新：{IC_MATRIX}")

    # 计算IR并排名
    rows = []
    recent3 = all_months[-2:] + [target_month] if len(all_months) >= 2 else [target_month]
    recent6 = all_months[-5:] + [target_month] if len(all_months) >= 5 else [target_month]

    for factor in ic_df.index:
        ic3 = [ic_df.loc[factor, m] for m in recent3 if m in ic_df.columns]
        ic6 = [ic_df.loc[factor, m] for m in recent6 if m in ic_df.columns]
        rows.append({
            "因子": factor,
            "本月IC": ic_results.get(factor, np.nan),
            "近3月均IC": round(np.nanmean(ic3), 4) if ic3 else np.nan,
            "近3月IR": calc_ir(ic3),
            "近6月IR": calc_ir(ic6),
        })

    result_df = pd.DataFrame(rows)
    result_df["评分"] = (
        result_df["近3月均IC"].abs().fillna(0) * 0.5 +
        result_df["近3月IR"].abs().fillna(0) / (result_df["近3月IR"].abs().max() + 1e-9) * 0.5
    )
    result_df = result_df.sort_values("评分", ascending=False).reset_index(drop=True)
    result_df["排名"] = range(1, len(result_df)+1)

    # ─── 输出 Top N ───────────────────────────────────
    print(f"\\n{'='*60}")
    print(f"IC/IR 排名 TOP {args.top}（{target_month}）")
    print(f"{'='*60}")
    print(f"{'排名':<4} {'因子':<18} {'本月IC':<10} {'近3月均IC':<12} {'近3月IR':<10}")
    print("-"*60)
    for _, r in result_df.head(args.top).iterrows():
        status = "✅" if r["近3月均IC"] > 0.1 and r["近3月IR"] >= 1.0 else "⚠️" if r["近3月均IC"] > 0.05 else "❌"
        print(f"{int(r['排名']):<4} {r['因子']:<18} {r['本月IC']:<10.4f} {r['近3月均IC']:<12.4f} {r['近3月IR']:<10.2f} {status}")
    print(f"{'='*60}")

    result_df.to_csv(os.path.join(OUTPUT_DIR, "factor_ic_summary.csv"), index=False)
    print(f"\\n完整结果已保存: {OUTPUT_DIR}/factor_ic_summary.csv")
'''

CALC_PATH = '/Users/ydy/Downloads/calculate_ic.py'
with open(CALC_PATH, 'w', encoding='utf-8') as f:
    f.write(CALC_TEMPLATE)
print(f"✅ calculate_ic.py 已更新: {CALC_PATH}")
print("\n🎉 全部文件生成完毕！")
