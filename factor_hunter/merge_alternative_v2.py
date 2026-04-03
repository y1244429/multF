#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
合并另类因子到V5 Excel - 最终版
"""
import os, warnings
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
warnings.filterwarnings('ignore')

OUTPUT_DIR = '/Users/ydy/WorkBuddy/20260326134350/factor_hunter/output'
EXCEL_PATH = os.path.join(OUTPUT_DIR, '因子监控_IC_IR_实测版_v5.xlsx')

# ── 另类因子IC结果（来自alternative_factors_v2.py）────────────────────
ALPHA_IC = [
    # analyst_eps_2025 冠军
    {
        '因子名称': 'analyst_eps_2025', '大类': '另类数据-分析师情绪',
        '细类': '预测EPS绝对值', '典型代表': 'EPS>5元/股=大白马',
        '平均IC': 0.0656, 'IC方向': '正向↑', '|IC|': 0.0656,
        'IC>0比例': '70.4%', 'IR': 0.436, 'IC_std': 0.150,
        '月份数': 27, '论文来源': 'AKShare analyst_forecast_em',
        'IC稳定性★': '★★★☆ 较稳定(IR=0.44)', '换手率代理★': '★☆☆☆☆ 极低(季频)',
        '因子独立性★': '★★★★★ 极高(基本面)', '市场稳健性★': '★★★☆ 较稳健',
        '逻辑可持续性★': '★★★★★ 极强(EPS=盈利质量)',
        '综合评分': 3.9,
        '说明': '2025年预测EPS，IC=+0.066, IR=0.44, IC>0率70.4%。分析师预测EPS越高=基本面越好=Alpha越强。',
    },
    # analyst_attention
    {
        '因子名称': 'analyst_attention', '大类': '另类数据-分析师情绪',
        '细类': '分析师关注度', '典型代表': '研报数>20篇=机构高覆盖',
        '平均IC': 0.0440, 'IC方向': '正向↑', '|IC|': 0.0440,
        'IC>0比例': '51.9%', 'IR': 0.292, 'IC_std': 0.151,
        '月份数': 27, '论文来源': 'AKShare analyst_forecast_em',
        'IC稳定性★': '★★☆☆ 较稳定(IR=0.29)', '换手率代理★': '★☆☆☆☆ 极低(季频)',
        '因子独立性★': '★★★★☆ 高(分析师覆盖≠价量)', '市场稳健性★': '★★★☆ 较稳健',
        '逻辑可持续性★': '★★★★☆ 强(关注度=机构化)',
        '综合评分': 2.6,
        '说明': 'ln(研报数+1)，IC=+0.044, IR=0.29, IC>0率51.9%。分析师覆盖越多=机构化程度越高=流动性溢价。',
    },
    # analyst_buy_ratio
    {
        '因子名称': 'analyst_buy_ratio', '大类': '另类数据-分析师情绪',
        '细类': '买入比例', '典型代表': '买入/(买入+增持+中性)>70%',
        '平均IC': 0.0034, 'IC方向': '正向↑', '|IC|': 0.0034,
        'IC>0比例': '55.6%', 'IR': 0.031, 'IC_std': 0.111,
        '月份数': 27, '论文来源': 'AKShare analyst_forecast_em',
        'IC稳定性★': '★☆☆☆☆ 弱(IR=0.03)', '换手率代理★': '★☆☆☆☆ 极低(季频)',
        '因子独立性★': '★★★★☆ 高', '市场稳健性★': '★★☆☆ 弱(信号弱)',
        '逻辑可持续性★': '★★★☆☆ 中(A股分析师偏乐观)',
        '综合评分': 1.6,
        '说明': '买入评级占比，IC=+0.003, IR=0.03, IC>0率55.6%。信号较弱，A股分析师普遍乐观，评级同质化。',
    },
    # analyst_net_rating
    {
        '因子名称': 'analyst_net_rating', '大类': '另类数据-分析师情绪',
        '细类': '综合评级', '典型代表': '加权评级>4.5=强烈买入',
        '平均IC': 0.0039, 'IC方向': '正向↑', '|IC|': 0.0039,
        'IC>0比例': '55.6%', 'IR': 0.035, 'IC_std': 0.113,
        '月份数': 27, '论文来源': 'AKShare analyst_forecast_em',
        'IC稳定性★': '★☆☆☆☆ 弱(IR=0.04)', '换手率代理★': '★☆☆☆☆ 极低(季频)',
        '因子独立性★': '★★★★☆ 高', '市场稳健性★': '★★☆☆ 弱(信号弱)',
        '逻辑可持续性★': '★★★☆☆ 中',
        '综合评分': 1.7,
        '说明': '加权综合评级，IC=+0.004, IR=0.04, IC>0率55.6%。与buy_ratio类似，信号弱但方向正确。',
    },
    # analyst_eps_fwd_growth
    {
        '因子名称': 'analyst_eps_fwd_growth', '大类': '另类数据-分析师情绪',
        '细类': '预测EPS调整', '典型代表': '(2025EPS-2024EPS)/|2024EPS|>10%',
        '平均IC': 0.0066, 'IC方向': '正向↑', '|IC|': 0.0066,
        'IC>0比例': '48.1%', 'IR': 0.078, 'IC_std': 0.085,
        '月份数': 27, '论文来源': 'AKShare analyst_forecast_em',
        'IC稳定性★': '★☆☆☆☆ 弱(IR=0.08)', '换手率代理★': '★☆☆☆☆ 极低(季频)',
        '因子独立性★': '★★★★★ 极高', '市场稳健性★': '★★☆☆ 弱(IC<50%)',
        '逻辑可持续性★': '★★★☆☆ 中(预测上调≠股价涨)',
        '综合评分': 1.2,
        '说明': 'EPS预测调整，IC=+0.007, IR=0.08, IC>0率48.1%。分析师上调预测≠股价涨（分析师往往左侧唱多）。',
    },
    # weibo_sentiment（微博情绪 - 单期IC）
    {
        '因子名称': 'weibo_sentiment', '大类': '另类数据-微博情绪',
        '细类': '微博舆情', '典型代表': '微博rate>0=正面舆情',
        '平均IC': 0.0446, 'IC方向': '正向↑', '|IC|': 0.0446,
        'IC>0比例': '单期', 'IR': '待滚动', 'IC_std': '单期',
        '月份数': 1, '论文来源': 'AKShare stock_js_weibo_report',
        'IC稳定性★': '? 需滚动验证', '换手率代理★': '★☆☆☆☆ 极低(周频更新)',
        '因子独立性★': '★★★★★ 极高(社媒情绪)', '市场稳健性★': '★★☆☆ 仅50只热门股',
        '逻辑可持续性★': '★★★☆☆ 中(舆情≠基本面)',
        '综合评分': 2.0,
        '说明': '微博近12小时舆情情绪，单期IC=+0.045。覆盖50只热门股，舆情正面=关注度高=资金流入。',
    },
    # pmi_manufacturing（PMI宏观）
    {
        '因子名称': 'pmi_manufacturing', '大类': '另类数据-PMI宏观',
        '细类': 'PMI宏观代理', '典型代表': 'PMI>50=扩张期',
        '平均IC': -0.103, 'IC方向': '逆向↓', '|IC|': 0.103,
        'IC>0比例': '宏观-全市场同向', 'IR': '宏观无截面IC', 'IC_std': '宏观',
        '月份数': 1, '论文来源': '国家统计局PMI',
        'IC稳定性★': '★★★☆ 宏观有效性', '换手率代理★': '★☆☆☆☆ 月度(宏观不变)',
        '因子独立性★': '★★★☆☆ 宏观(所有股同向)', '市场稳健性★': '★★★★☆ 宏观逻辑强',
        '逻辑可持续性★': '★★★★☆ 强(PMI→经济预期)',
        '综合评分': 2.0,
        '说明': '制造业PMI标准化，所有股票同时暴露。PMI下降期(PMI<50)=经济收缩=股市跌，逆向因子。',
    },
]

# ── 通达信另类因子公式 ───────────────────────────────────────────────
TONGDAXIN_FORMULAS = [
    ('另-A公式（推荐）', '分析师EPS预测因子', 'analyst_eps_2025 IC=+0.066', '另-A'),
    ('另-B公式', '分析师关注度因子', 'analyst_attention IC=+0.044', '另-B'),
    ('另-C公式', '微博舆情因子', 'weibo_sentiment IC=+0.045', '另-C'),
]

def main():
    print("合并另类因子到V5 Excel...")
    wb = openpyxl.load_workbook(EXCEL_PATH)
    print(f"  Sheets: {wb.sheetnames}")

    # ── 1. 因子监控总表 ──
    ws1 = wb['因子监控总表']
    # 找最后一行
    last_row = ws1.max_row
    while last_row > 3 and ws1.cell(last_row, 1).value in (None, ''):
        last_row -= 1

    fills = {
        '另类数据-分析师情绪': PatternFill('solid', fgColor='FFF2CC'),
        '另类数据-微博情绪': PatternFill('solid', fgColor='E2EFDA'),
        '另类数据-PMI宏观': PatternFill('solid', fgColor='DEEBF7'),
    }

    # 表头行（第3行）
    header_row = 3
    headers = [ws1.cell(header_row, c).value for c in range(1, ws1.max_column + 1)]

    for row_data in ALPHA_IC:
        last_row += 2
        for ci, col_name in enumerate(headers, 1):
            val = row_data.get(col_name, '')
            if isinstance(val, float) and np.isnan(val):
                val = ''
            cell = ws1.cell(last_row, ci, val)
            cat = row_data.get('大类', '')
            if cat in fills:
                cell.fill = fills[cat]

    print(f"  因子监控总表: 追加 {len(ALPHA_IC)} 行, 到第{last_row}行")

    # ── 2. 月度IC矩阵 ──
    if '月度IC矩阵' in wb.sheetnames:
        ws2 = wb['月度IC矩阵']
        print(f"  月度IC矩阵: {ws2.max_row} 行")

    # ── 3. 论文来源说明 ──
    if '论文来源说明' in wb.sheetnames:
        ws3 = wb['论文来源说明']
        lr = ws3.max_row + 2
        ws3.cell(lr, 1, '── 另类数据因子（2026-04-02）──')
        ws3.cell(lr, 1).font = Font(bold=True, size=12, color='FF6600')
        lr += 1
        for item in ALPHA_IC:
            lr += 1
            ws3.cell(lr, 1, f"[{item['因子名称']}] {item['说明']}")
            ws3.cell(lr, 2, f"IC={item['平均IC']:+.4f}, IR={item['IR']}")
        print(f"  论文来源说明: 已追加")

    # ── 4. 通达信选股公式 ──
    ws4 = wb['通达信选股公式']
    lr4 = ws4.max_row + 2
    ws4.cell(lr4, 1, '── 另类数据因子选股公式（2026-04-02）──')
    ws4.cell(lr4, 1).font = Font(bold=True, size=12, color='FF6600')
    lr4 += 1

    formulas = [
        ('另-A公式（推荐）',
         '分析师EPS预测 + 综合评级 + 关注度',
         ('{条件1}2025预测EPS>2元, 要求: 必须满足(基本面优质)\\n'
          '{条件2}综合评级>4分(买入+增持>50%), 要求: 必须满足\\n'
          '{条件3}研报数>5篇, 要求: 必须满足(分析师覆盖)\\n'
          '{说明}IC=+0.066, IR=0.44, IC>0率70.4%——分析师预测EPS越高=基本面越好=Alpha'),
         'analyst_eps_2025 IC=+0.066(滚动27月冠军), analyst_attention IC=+0.044, analyst_net_rating IC=+0.004',
         '另-A'),
        ('另-B公式（微博舆情）',
         '微博正面舆情 + 有热度覆盖',
         ('{条件1}微博舆情rate>0, 要求: 必须满足(仅50只热门股)\\n'
          '{条件2}有微博热度, 要求: 必须满足\\n'
          '{说明}IC=+0.045(单期)，微博正面=资金关注=短期动量，但仅覆盖热门股'),
         'weibo_sentiment IC=+0.045(单期), weibo_buzz IC=+0.045',
         '另-B'),
        ('另-C公式（PMI宏观择时）',
         'PMI扩张期 + 制造业指数',
         ('{说明}宏观择时公式，不适用于选股。适用于大盘择时/仓位管理。\\n'
          'PMI>50(制造业扩张期) → 仓位偏多\\n'
          'PMI<50(制造业收缩期) → 仓位偏空\\n'
          '当前PMI=-0.103(标准化)，需根据实际PMI数值调整阈值'),
         'pmi_manufacturing: PMI>50扩张做多, PMI<50收缩减仓',
         '另-C'),
    ]
    for fname, ftype, fdesc, flogic, ftag in formulas:
        lr4 += 1
        ws4.cell(lr4, 1, fname)
        ws4.cell(lr4, 2, ftype)
        ws4.cell(lr4, 3, fdesc)
        ws4.cell(lr4, 4, flogic)
        ws4.cell(lr4, 5, ftag)
        # 高亮
        for ci in range(1, 6):
            ws4.cell(lr4, ci).fill = PatternFill('solid', fgColor='FFF2CC')

    # 保存
    wb.save(EXCEL_PATH)
    print(f"  ✅ Excel已保存: {EXCEL_PATH}")

    # 汇总
    print("\n" + "=" * 60)
    print("另类因子最终汇总（27个月滚动IC）")
    print("=" * 60)
    print(f"{'因子':25s} {'IC':>8s} {'IR':>6s} {'IC>0率':>8s} {'评分':>5s}")
    print("-" * 60)
    for item in sorted(ALPHA_IC, key=lambda x: x['综合评分'], reverse=True):
        print(f"{item['因子名称']:25s} {item['平均IC']:>+8.4f} {item['IR']:>6} {item['IC>0比例']:>8s} {item['综合评分']:>5.1f}")
    print("=" * 60)

if __name__ == '__main__':
    main()
